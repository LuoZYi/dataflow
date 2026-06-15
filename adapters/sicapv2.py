# dataflow/adapters/sicapv2.py
from __future__ import annotations

import random
from pathlib import Path
from typing import Any, Dict, Iterator, List, Optional, Set, Tuple

import numpy as np
from PIL import Image

from .types import AnnObject, BaseAdapter, Sample
from .utils import (
    index_by_stem,
    read_rgb,
    bbox_from_mask,
    area_from_mask,
    read_hw_fast,
)


def _clean_stem(x: Any) -> Optional[str]:
    if x is None:
        return None

    s = str(x).strip()
    if not s:
        return None

    return Path(Path(s).name).stem


def _read_sicap_partition_xlsx(path: Path) -> Tuple[Set[str], Dict[str, Dict[str, int]]]:
    """
    Read SICAPv2 partition xlsx.

    Expected columns:
      image_name, NC, G3, G4, G5, G4C
    """
    stems: Set[str] = set()
    labels: Dict[str, Dict[str, int]] = {}

    if not path.exists():
        return stems, labels

    try:
        from openpyxl import load_workbook
    except Exception as e:
        raise RuntimeError(f"[SICAPv2] openpyxl is required: {path}") from e

    wb = load_workbook(path, read_only=True, data_only=True)

    for ws in wb.worksheets:
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)

        if header is None:
            continue

        header_names = [str(x).strip() if x is not None else "" for x in header]
        header_lower = [x.lower() for x in header_names]

        image_col = header_lower.index("image_name") if "image_name" in header_lower else 0

        label_cols: Dict[str, int] = {}
        for key in ["NC", "G3", "G4", "G5", "G4C"]:
            if key.lower() in header_lower:
                label_cols[key] = header_lower.index(key.lower())

        for row in rows:
            if not row or len(row) <= image_col:
                continue

            stem = _clean_stem(row[image_col])
            if stem is None:
                continue

            stems.add(stem)

            cur: Dict[str, int] = {}
            for key, idx in label_cols.items():
                if len(row) <= idx:
                    cur[key] = 0
                    continue
                try:
                    cur[key] = int(row[idx]) if row[idx] is not None else 0
                except Exception:
                    cur[key] = 0

            labels[stem] = cur

    return stems, labels


def _decode_sicap_gray_mask(mask_path: Path) -> np.ndarray:
    gray = np.asarray(Image.open(mask_path).convert("L"))

    out = np.zeros(gray.shape, dtype=np.uint8)

    # JPEG-corrupted encoded labels:
    # 0   -> non-cancerous/background
    # 100 -> GG3
    # 150 -> GG4
    # 200 -> GG5
    out[(gray >= 50) & (gray < 125)] = 1
    out[(gray >= 125) & (gray < 175)] = 2
    out[gray >= 175] = 3

    return out


class SICAPv2Adapter(BaseAdapter):
    dataset_name = "sicapv2"

    def __init__(
        self,
        root: Path,
        *,
        images_dir: str = "images",
        masks_dir: str = "masks",
        partition_dir: str = "partition",
        split_mode: str = "official",  # "official" | "random" | "unspecified"
        val_fold: int = 1,
        seed: int = 42,
        val_ratio: float = 0.1,
        test_ratio: float = 0.1,
        skip_unlisted: bool = True,
        min_area: int = 64,
    ) -> None:
        super().__init__(root)

        if (self.root / images_dir).exists() and (self.root / masks_dir).exists():
            self.data_root = self.root
        elif (
            (self.root / "SICAPv2" / images_dir).exists()
            and (self.root / "SICAPv2" / masks_dir).exists()
        ):
            self.data_root = self.root / "SICAPv2"
        else:
            self.data_root = self.root

        self.images_dir = images_dir
        self.masks_dir = masks_dir
        self.partition_dir = partition_dir
        self.split_mode = split_mode
        self.val_fold = int(val_fold)
        self.seed = seed
        self.val_ratio = val_ratio
        self.test_ratio = test_ratio
        self.skip_unlisted = skip_unlisted
        self.min_area = min_area

        if self.val_fold not in {1, 2, 3, 4}:
            raise ValueError(f"[SICAPv2] val_fold must be 1, 2, 3, or 4, got {val_fold}")

        # Pixel-level mask labels after quantization.
        self.id2name: Dict[int, str] = {
            1: "gleason_grade_3",
            2: "gleason_grade_4",
            3: "gleason_grade_5",
        }

        self._xlsx_labels: Dict[str, Dict[str, int]] = {}
        self._last_split_debug: Dict[str, int] = {}

    def _assign_random_splits(self, sample_ids: List[str]) -> Dict[str, str]:
        rng = random.Random(self.seed)
        ids = list(sample_ids)
        rng.shuffle(ids)

        n = len(ids)
        n_test = int(round(n * self.test_ratio))
        n_val = int(round(n * self.val_ratio))

        test = set(ids[:n_test])
        val = set(ids[n_test : n_test + n_val])

        out: Dict[str, str] = {}
        for sid in ids:
            if sid in test:
                out[sid] = "test"
            elif sid in val:
                out[sid] = "val"
            else:
                out[sid] = "train"

        return out

    def _assign_official_splits(self, sample_ids: List[str]) -> Dict[str, str]:
        sample_set = set(sample_ids)

        test_xlsx = self.data_root / self.partition_dir / "Test" / "Test.xlsx"

        val_dir = (
            self.data_root
            / self.partition_dir
            / "Validation"
            / f"Val{self.val_fold}"
        )

        val_xlsx = val_dir / "Test.xlsx"
        train_xlsx = val_dir / "Train.xlsx"

        test_raw, test_labels = _read_sicap_partition_xlsx(test_xlsx)
        val_raw, val_labels = _read_sicap_partition_xlsx(val_xlsx)
        train_raw, train_labels = _read_sicap_partition_xlsx(train_xlsx)

        self._xlsx_labels = {}
        self._xlsx_labels.update(train_labels)
        self._xlsx_labels.update(val_labels)
        self._xlsx_labels.update(test_labels)

        test_ids = test_raw & sample_set
        val_ids = val_raw & sample_set
        train_ids = train_raw & sample_set

        val_ids = val_ids - test_ids
        train_ids = train_ids - test_ids - val_ids

        out: Dict[str, str] = {}

        for sid in sample_ids:
            if sid in test_ids:
                out[sid] = "test"
            elif sid in val_ids:
                out[sid] = "val"
            elif sid in train_ids:
                out[sid] = "train"
            else:
                out[sid] = "ignore" if self.skip_unlisted else "train"

        self._last_split_debug = {
            "total_image_mask_pairs": len(sample_ids),
            "xlsx_train_matched": len(train_ids),
            "xlsx_val_matched": len(val_ids),
            "xlsx_test_matched": len(test_ids),
            "ignore": sum(1 for x in out.values() if x == "ignore"),
            "train": sum(1 for x in out.values() if x == "train"),
            "val": sum(1 for x in out.values() if x == "val"),
            "test": sum(1 for x in out.values() if x == "test"),
            "test_xlsx_exists": int(test_xlsx.exists()),
            "val_xlsx_exists": int(val_xlsx.exists()),
            "train_xlsx_exists": int(train_xlsx.exists()),
        }

        return out

    def _assign_splits(self, sample_ids: List[str]) -> Dict[str, str]:
        if self.split_mode == "unspecified":
            return {sid: "unspecified" for sid in sample_ids}

        if self.split_mode == "random":
            return self._assign_random_splits(sample_ids)

        if self.split_mode == "official":
            return self._assign_official_splits(sample_ids)

        raise ValueError(f"[SICAPv2] Unknown split_mode={self.split_mode}")

    def _find_mask_for_image(self, img_path: Path, mask_map: Dict[str, Path]) -> Optional[Path]:
        stem = img_path.stem

        if stem in mask_map:
            return mask_map[stem]

        for cand in [
            stem + "_mask",
            stem + "-mask",
            stem + "_label",
            stem + "-label",
            stem + "_gt",
            stem + "-gt",
        ]:
            if cand in mask_map:
                return mask_map[cand]

        msk_dir = self.data_root / self.masks_dir
        hits: List[Path] = []
        for ext in [".jpg", ".jpeg", ".png", ".tif", ".tiff"]:
            hits.extend(msk_dir.glob(stem + "*" + ext))

        return sorted(hits)[0] if hits else None

    def iter_samples(self) -> Iterator[Sample]:
        img_dir = self.data_root / self.images_dir
        msk_dir = self.data_root / self.masks_dir

        if not img_dir.exists():
            raise FileNotFoundError(f"[SICAPv2] images dir not found: {img_dir}")

        if not msk_dir.exists():
            raise FileNotFoundError(f"[SICAPv2] masks dir not found: {msk_dir}")

        img_map = index_by_stem(img_dir)
        msk_map = index_by_stem(msk_dir)

        records: List[Tuple[str, Path, Path]] = []
        sample_ids: List[str] = []

        for stem, img_path in sorted(img_map.items()):
            mask_path = self._find_mask_for_image(img_path, msk_map)
            if mask_path is None:
                continue

            sample_ids.append(stem)
            records.append((stem, img_path, mask_path))

        if not records:
            raise ValueError(
                f"[SICAPv2] No matched images/masks under {img_dir} and {msk_dir}"
            )

        split_of = self._assign_splits(sample_ids)

        for sample_id, img_path, mask_path in records:
            split = split_of.get(sample_id, "ignore")
            if split == "ignore":
                continue

            H, W = read_hw_fast(str(img_path))
            case_id = sample_id.split("_Block_")[0] if "_Block_" in sample_id else sample_id

            yield Sample(
                dataset=self.dataset_name,
                sample_id=sample_id,
                split=split,  # type: ignore
                image_path=img_path,
                ann_path=mask_path,
                group_id=case_id,
                meta={
                    "height_px": H,
                    "width_px": W,
                    "case_id": case_id,
                    "label_map": dict(self.id2name),
                    "xlsx_labels": dict(self._xlsx_labels.get(sample_id, {})),
                    "data_root": str(self.data_root),
                    "val_fold": self.val_fold,
                    "split_mode": self.split_mode,
                    "skip_unlisted": self.skip_unlisted,
                    "mask_encoding": "jpeg_gray_quantized_0_100_150_200",
                },
            )

    def load_image(self, sample: Sample) -> np.ndarray:
        return read_rgb(sample.image_path)

    def iter_ann(self, sample: Sample) -> Iterator[AnnObject]:
        assert sample.ann_path is not None

        label_map = _decode_sicap_gray_mask(sample.ann_path)

        for cid in sorted(int(x) for x in np.unique(label_map).tolist() if int(x) != 0):
            m = label_map == cid
            area = area_from_mask(m)

            if area < self.min_area:
                continue

            yield AnnObject(
                ann_id=f"{sample.sample_id}:c{cid}",
                kind="semantic",
                source_label=self.id2name.get(cid, f"class_{cid}"),
                source_label_id=cid,
                mask=m,
                bbox_xywh=bbox_from_mask(m),
                area=area,
            )